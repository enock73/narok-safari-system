import os
import tempfile
from datetime import datetime
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                  Paragraph, Spacer, HRFlowable)
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app import db
from app.models import GateClearance, Revenue, WildlifeSighting, User, Vehicle
from sqlalchemy import func

DARK_GREEN = colors.HexColor('#1a3c2e')
LIGHT_GREEN = colors.HexColor('#2d6a4f')
ACCENT = colors.HexColor('#f4a261')
LIGHT_BG = colors.HexColor('#f8fdf9')


def generate_pdf_report(report_type, date_from, date_to):
    """Generate a PDF report and return the file path."""
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    tmp.close()

    doc = SimpleDocTemplate(
        tmp.name,
        pagesize=landscape(A4),
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=18,
        textColor=DARK_GREEN,
        spaceAfter=6,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=LIGHT_GREEN,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    normal_style = ParagraphStyle(
        'Body',
        parent=styles['Normal'],
        fontSize=9
    )

    story = []
    story.append(Paragraph('NAROK COUNTY GOVERNMENT', title_style))
    story.append(Paragraph('Maasai Mara Ecosystem Management & Safari Gate Clearance System', subtitle_style))
    story.append(HRFlowable(width='100%', thickness=2, color=DARK_GREEN, spaceAfter=10))

    if report_type == 'clearances':
        story += _clearances_pdf_content(date_from, date_to, normal_style, styles)
    elif report_type == 'revenue':
        story += _revenue_pdf_content(date_from, date_to, normal_style, styles)
    elif report_type == 'wildlife':
        story += _wildlife_pdf_content(date_from, date_to, normal_style, styles)
    elif report_type == 'guides':
        story += _guides_pdf_content(normal_style, styles)

    story.append(Spacer(1, 20))
    story.append(Paragraph(
        f'Report generated: {datetime.now().strftime("%d %B %Y at %H:%M")} | Narok County Government',
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=7,
                       textColor=colors.grey, alignment=TA_CENTER)
    ))

    doc.build(story)
    return tmp.name


def _table_style(header_bg=DARK_GREEN):
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), header_bg),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_BG]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ])


def _clearances_pdf_content(date_from, date_to, normal_style, styles):
    story = []
    story.append(Paragraph(
        f'Gate Clearance Report  |  {date_from} to {date_to}',
        ParagraphStyle('SectionTitle', parent=styles['Heading2'],
                       textColor=DARK_GREEN, fontSize=13, spaceAfter=10)
    ))

    clearances = GateClearance.query.filter(
        GateClearance.entry_date >= date_from,
        GateClearance.entry_date <= date_to
    ).order_by(GateClearance.entry_date).all()

    # Summary
    approved = sum(1 for c in clearances if c.status == 'approved')
    total_pax = sum(c.passenger_count or 0 for c in clearances)
    story.append(Paragraph(
        f'Total Clearances: <b>{len(clearances)}</b> &nbsp;|&nbsp; '
        f'Approved: <b>{approved}</b> &nbsp;|&nbsp; '
        f'Total Passengers: <b>{total_pax}</b>',
        normal_style
    ))
    story.append(Spacer(1, 10))

    headers = ['#', 'Token', 'Guide', 'Vehicle', 'Gate', 'Date', 'Pax', 'Status', 'Fee (KES)']
    data = [headers]
    for i, c in enumerate(clearances, 1):
        data.append([
            str(i),
            c.token[:8] + '...',
            c.guide.full_name[:20],
            c.vehicle.plate_number,
            c.gate.replace(' Gate', ''),
            str(c.entry_date),
            str(c.passenger_count or 0),
            c.status.upper(),
            f'{float(c.fee_paid or 0):,.0f}'
        ])

    if len(data) > 1:
        col_widths = [1*cm, 2.5*cm, 4*cm, 3*cm, 4*cm, 2.5*cm, 1.5*cm, 2.5*cm, 2.5*cm]
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(_table_style())
        story.append(t)
    else:
        story.append(Paragraph('No clearances in this period.', normal_style))

    return story


def _revenue_pdf_content(date_from, date_to, normal_style, styles):
    story = []
    story.append(Paragraph(
        f'Revenue Report  |  {date_from} to {date_to}',
        ParagraphStyle('SectionTitle', parent=styles['Heading2'],
                       textColor=DARK_GREEN, fontSize=13, spaceAfter=10)
    ))

    revenues = Revenue.query.filter(
        func.date(Revenue.collected_at) >= date_from,
        func.date(Revenue.collected_at) <= date_to
    ).all()

    total = sum(float(r.amount) for r in revenues)
    story.append(Paragraph(
        f'Total Revenue: <b>KES {total:,.2f}</b>  |  Transactions: <b>{len(revenues)}</b>',
        normal_style
    ))
    story.append(Spacer(1, 10))

    # By gate
    by_gate = {}
    for r in revenues:
        by_gate[r.gate] = by_gate.get(r.gate, 0) + float(r.amount)

    gate_data = [['Gate', 'Amount (KES)', 'Share %']]
    for gate, amt in sorted(by_gate.items(), key=lambda x: -x[1]):
        pct = (amt / total * 100) if total else 0
        gate_data.append([gate, f'{amt:,.2f}', f'{pct:.1f}%'])

    if len(gate_data) > 1:
        t = Table(gate_data, colWidths=[7*cm, 5*cm, 4*cm])
        t.setStyle(_table_style())
        story.append(t)
        story.append(Spacer(1, 15))

    headers = ['Date', 'Gate', 'Method', 'Amount (KES)', 'Clearance Token']
    data = [headers]
    for r in revenues:
        data.append([
            r.collected_at.strftime('%Y-%m-%d %H:%M'),
            r.gate or '',
            r.payment_method or 'Cash',
            f'{float(r.amount):,.2f}',
            r.clearance.token[:8] + '...' if r.clearance else ''
        ])

    if len(data) > 1:
        col_widths = [4*cm, 5*cm, 4*cm, 4*cm, 4*cm]
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(_table_style())
        story.append(t)

    return story


def _wildlife_pdf_content(date_from, date_to, normal_style, styles):
    story = []
    story.append(Paragraph(
        f'Wildlife Sightings Report  |  {date_from} to {date_to}',
        ParagraphStyle('SectionTitle', parent=styles['Heading2'],
                       textColor=DARK_GREEN, fontSize=13, spaceAfter=10)
    ))

    sightings = WildlifeSighting.query.filter(
        func.date(WildlifeSighting.sighted_at) >= date_from,
        func.date(WildlifeSighting.sighted_at) <= date_to
    ).order_by(WildlifeSighting.sighted_at.desc()).all()

    total_animals = sum(s.count for s in sightings)
    story.append(Paragraph(
        f'Total Reports: <b>{len(sightings)}</b>  |  Total Animals: <b>{total_animals}</b>',
        normal_style
    ))
    story.append(Spacer(1, 10))

    headers = ['Species', 'Category', 'Count', 'Location', 'Behavior', 'Date/Time', 'Threat']
    data = [headers]
    for s in sightings:
        data.append([
            (s.common_name or s.species)[:20],
            s.category,
            str(s.count),
            (s.location_name or f'{float(s.latitude):.4f},{float(s.longitude):.4f}')[:25],
            s.behavior or '',
            s.sighted_at.strftime('%Y-%m-%d %H:%M'),
            s.threat_level.upper()
        ])

    if len(data) > 1:
        col_widths = [4*cm, 3*cm, 1.5*cm, 5*cm, 2.5*cm, 4*cm, 2*cm]
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(_table_style())
        story.append(t)
    else:
        story.append(Paragraph('No sightings in this period.', normal_style))

    return story


def _guides_pdf_content(normal_style, styles):
    story = []
    story.append(Paragraph(
        'Registered Tour Guides',
        ParagraphStyle('SectionTitle', parent=styles['Heading2'],
                       textColor=DARK_GREEN, fontSize=13, spaceAfter=10)
    ))

    guides = User.query.filter_by(role='guide').order_by(User.full_name).all()
    headers = ['Name', 'Email', 'Phone', 'Company', 'License', 'Verified', 'Clearances']
    data = [headers]
    for g in guides:
        data.append([
            g.full_name[:25],
            g.email[:25],
            g.phone or '-',
            (g.company or '-')[:20],
            g.license_number or '-',
            'Yes' if g.is_verified else 'No',
            str(g.clearances.count())
        ])

    if len(data) > 1:
        col_widths = [4*cm, 5*cm, 3*cm, 4*cm, 3*cm, 2*cm, 2.5*cm]
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(_table_style())
        story.append(t)

    return story


# ── Excel Reports ─────────────────────────────────────────────────────────────

def generate_excel_report(report_type, date_from, date_to):
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    tmp.close()

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1a3c2e')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alt_fill = PatternFill('solid', fgColor='f8fdf9')
    border = Border(
        left=Side(style='thin', color='cccccc'),
        right=Side(style='thin', color='cccccc'),
        top=Side(style='thin', color='cccccc'),
        bottom=Side(style='thin', color='cccccc')
    )
    title_font = Font(bold=True, color='1a3c2e', size=14)

    def style_header(ws, headers, row=3):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

    def add_title(ws, title, subtitle):
        ws.merge_cells('A1:J1')
        ws['A1'] = 'NAROK COUNTY GOVERNMENT — Maasai Mara Ecosystem Management'
        ws['A1'].font = Font(bold=True, color='1a3c2e', size=13)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:J2')
        ws['A2'] = f'{title}  |  {subtitle}'
        ws['A2'].font = Font(color='2d6a4f', size=11)
        ws['A2'].alignment = Alignment(horizontal='center')

    if report_type == 'clearances':
        ws = wb.active
        ws.title = 'Gate Clearances'
        add_title(ws, 'Gate Clearance Report', f'{date_from} to {date_to}')
        headers = ['#', 'Token', 'Guide Name', 'Company', 'Vehicle', 'Gate',
                   'Entry Date', 'Passengers', 'Purpose', 'Status', 'Fee (KES)']
        style_header(ws, headers)

        clearances = GateClearance.query.filter(
            GateClearance.entry_date >= date_from,
            GateClearance.entry_date <= date_to
        ).all()

        for i, c in enumerate(clearances, 1):
            row = i + 3
            row_data = [
                i, c.token, c.guide.full_name, c.guide.company or '',
                c.vehicle.plate_number, c.gate, str(c.entry_date),
                c.passenger_count or 0, c.purpose, c.status.upper(),
                float(c.fee_paid or 0)
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = border
                if i % 2 == 0:
                    cell.fill = alt_fill

        # Auto width
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 35)

    elif report_type == 'revenue':
        ws = wb.active
        ws.title = 'Revenue'
        add_title(ws, 'Revenue Report', f'{date_from} to {date_to}')
        headers = ['#', 'Date', 'Gate', 'Guide', 'Amount (KES)',
                   'Currency', 'Method', 'M-Pesa Ref', 'Clearance Token']
        style_header(ws, headers)

        revenues = Revenue.query.filter(
            func.date(Revenue.collected_at) >= date_from,
            func.date(Revenue.collected_at) <= date_to
        ).all()

        for i, r in enumerate(revenues, 1):
            row = i + 3
            row_data = [
                i,
                r.collected_at.strftime('%Y-%m-%d %H:%M'),
                r.gate or '',
                r.clearance.guide.full_name if r.clearance else '',
                float(r.amount),
                r.currency,
                r.payment_method or 'Cash',
                r.mpesa_ref or '',
                r.clearance.token if r.clearance else ''
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = border
                if i % 2 == 0:
                    cell.fill = alt_fill

        # Summary sheet
        ws2 = wb.create_sheet('Summary by Gate')
        ws2['A1'] = 'Revenue by Gate'
        ws2['A1'].font = title_font
        ws2.append([])
        ws2.append(['Gate', 'Total (KES)', 'Transactions'])
        by_gate = db.session.query(
            Revenue.gate,
            func.sum(Revenue.amount).label('total'),
            func.count(Revenue.id).label('count')
        ).filter(
            func.date(Revenue.collected_at) >= date_from,
            func.date(Revenue.collected_at) <= date_to
        ).group_by(Revenue.gate).all()
        for g in by_gate:
            ws2.append([g.gate, float(g.total), g.count])

        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 35)

    elif report_type == 'wildlife':
        ws = wb.active
        ws.title = 'Wildlife Sightings'
        add_title(ws, 'Wildlife Sightings Report', f'{date_from} to {date_to}')
        headers = ['#', 'Species', 'Common Name', 'Category', 'Count',
                   'Latitude', 'Longitude', 'Location', 'Behavior',
                   'Date/Time', 'Threat Level', 'Verified']
        style_header(ws, headers)

        sightings = WildlifeSighting.query.filter(
            func.date(WildlifeSighting.sighted_at) >= date_from,
            func.date(WildlifeSighting.sighted_at) <= date_to
        ).all()

        for i, s in enumerate(sightings, 1):
            row = i + 3
            row_data = [
                i, s.species, s.common_name or '', s.category, s.count,
                float(s.latitude), float(s.longitude),
                s.location_name or '', s.behavior or '',
                s.sighted_at.strftime('%Y-%m-%d %H:%M'),
                s.threat_level, 'Yes' if s.is_verified else 'No'
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = border
                if i % 2 == 0:
                    cell.fill = alt_fill

        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 35)

    wb.save(tmp.name)
    return tmp.name
